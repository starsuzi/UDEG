from nltk import word_tokenize, pos_tag
import os
from pprint import pprint
from nltk.corpus import stopwords
from openpyxl import Workbook
from openpyxl.styles import Font
from nltk.stem import WordNetLemmatizer

lemmatizer = WordNetLemmatizer()

stopwords = stopwords.words("english") + [
    "``",
    "''",
    "?",
    ".",
    ",",
    "!",
    "'s",
]
path = "./sample/"
document_fname = os.path.join(path, "docs_test_text.txt")
query_fname = os.path.join(path, "queries_test_text.txt")
qrels_fname = os.path.join(path, "qrels_test_text.txt")
beam_result_fname = os.path.join(path, "test_pegasus_xsum_beam8_minlen3.txt")
mc4_result_fname = os.path.join(
    path, "test_pegasus_xsum_beam8_minlen5_4mc.txt"
)
mc5_result_fname = os.path.join(
    path, "test_pegasus_xsum_beam8_minlen5_5mc.txt"
)


def read_file(fname):
    with open(fname, "r") as f:
        return [el.strip().split("\t") for el in f.readlines()]


def read_qrels_file(fname):
    with open(fname, "r") as f:
        res = [el.strip().split("\t") for el in f.readlines()]
    # Return only relevant query-document pairs
    return [[el[0], el[2]] for el in res if int(el[3]) in [3, 4]]


def get_relevant_doc_list(qrels):
    result = {}
    for (qid, docid) in qrels:
        if qid not in result:
            result[qid] = []
        result[qid].append(docid)
    return result


def get_matched_dataset():
    """Query, document, generated text를 연결하고 tokenize해서 return합니다."""
    query_data = {el[0]: el[1] for el in read_file(query_fname)}
    document_data = {
        el[0]: {"doc_idx": idx, "text": el[1]}
        for idx, el in enumerate(read_file(document_fname))
    }
    qrels_data = read_qrels_file(qrels_fname)
    mc4_result_data = read_file(mc4_result_fname)

    packed_item = []
    for idx, (query_id, document_id) in enumerate(qrels_data):
        query = [el.lower() for el in word_tokenize(query_data[query_id])]
        document = [
            el.lower()
            for el in word_tokenize(document_data[document_id]["text"])
        ]
        document_idx = document_data[document_id]["doc_idx"]
        generated = mc4_result_data[document_idx]
        assert len(generated) == 1
        generated = [el.lower() for el in word_tokenize(generated[0])]
        packed_item.append(
            {
                "query": query,
                "document": document,
                "generated": generated,
                "query_id": query_id,
                "doc_id": document_id,
            }
        )
    return packed_item


def get_jaccard_index(a, b):
    """
    두 셋 사이의 유사도를 측정
    """
    return len(set(a).intersection(set(b))) / (
        len(a) + len(b) - len(set(a).intersection(set(b)))
    )


def get_score_data():
    with open("0315_data/baseline/run.baseline.qld.results.mrr", "r") as f:
        baseline = [el.strip().split() for el in f.readlines()]
    with open(
        "0315_data/4mc/run.pegasus_xsum_beam8_minlen5_4mc.qld.results.mrr",
        "r",
    ) as f:
        pega = [el.strip().split() for el in f.readlines()]

    baseline = [[el[1], float(el[2])] for el in baseline]
    pega = [[el[1], float(el[2])] for el in pega]
    assert len(pega) == len(baseline)
    merged = []
    for el1, el2 in zip(baseline, pega):
        assert el1[0] == el2[0]
        merged.append(el1 + [el2[1]])
    return merged


def write_excel(sample_candidate):
    wb = Workbook()
    ws = wb.active

    for idx, el in enumerate(sample_candidate):
        ws["A" + str(8 * idx + 1)] = "Query"
        ws["A" + str(8 * idx + 1)].font = Font(bold=True)
        ws["A" + str(8 * idx + 2)] = "Doc"
        ws["A" + str(8 * idx + 2)].font = Font(bold=True)
        ws["A" + str(8 * idx + 3)] = "Gen"
        ws["A" + str(8 * idx + 3)].font = Font(bold=True)
        ws["A" + str(8 * idx + 4)] = "Novel"
        ws["A" + str(8 * idx + 4)].font = Font(bold=True)
        ws["A" + str(8 * idx + 5)] = "Qid"
        ws["A" + str(8 * idx + 5)].font = Font(bold=True)
        ws["A" + str(8 * idx + 6)] = "Did"
        ws["A" + str(8 * idx + 6)].font = Font(bold=True)
        ws["A" + str(8 * idx + 7)] = "IDX"
        ws["A" + str(8 * idx + 7)].font = Font(bold=True)
        ws["B" + str(8 * idx + 1)] = el["query"]
        ws["B" + str(8 * idx + 2)] = el["doc"]
        ws["B" + str(8 * idx + 3)] = el["generated"]
        ws["B" + str(8 * idx + 4)] = " ".join(el["novel"])
        ws["B" + str(8 * idx + 5)] = el["qid"]
        ws["B" + str(8 * idx + 6)] = el["did"]
        ws["B" + str(8 * idx + 7)] = idx

    wb.save("effective_sample.xlsx")


def get_rank_result():
    baseline_fname = "./0315_data/baseline/run.baseline.qld.txt"
    pegasus_fname = (
        "./0315_data/4mc/run.pegasus_xsum_beam8_minlen5_4mc.qld.txt"
    )
    with open(baseline_fname, "r") as f:
        baseline = [el.strip().split() for el in f.readlines()]
        baseline = [[el[0], el[2], el[3]] for el in baseline]
    with open(pegasus_fname, "r") as f:
        pegasus = [el.strip().split() for el in f.readlines()]
        pegasus = [[el[0], el[2], el[3]] for el in pegasus]
    baseline_rank, pegasus_rank = {}, {}
    for el in baseline:
        if el[0] not in baseline_rank:
            baseline_rank[el[0]] = []
        baseline_rank[el[0]].append(el[1])
    for el in pegasus:
        if el[0] not in pegasus_rank:
            pegasus_rank[el[0]] = []
        pegasus_rank[el[0]].append(el[1])
    return baseline_rank, pegasus_rank


from tqdm import tqdm


def main():
    baseline_rank_map, pegasus_rank_map = get_rank_result()
    score_data = get_score_data()
    query_data = {el[0]: el[1] for el in read_file(query_fname)}
    document_data = {
        el[0]: {"doc_idx": idx, "text": el[1]}
        for idx, el in enumerate(read_file(document_fname))
    }
    qrels_data = read_qrels_file(qrels_fname)
    relevant_docs = get_relevant_doc_list(qrels_data)
    mc4_result_data = read_file(mc4_result_fname)

    final_result = []
    for (query_id, original_score, expanded_score) in tqdm(score_data):
        if query_id == "all":
            break
        query = query_data[query_id]
        baseline_rank_list, pegasus_rank_list = (
            baseline_rank_map[query_id],
            pegasus_rank_map[query_id],
        )

        relevant_document_list = relevant_docs[query_id]

        for relevant_document_id in relevant_document_list:
            try:
                assert relevant_document_id in baseline_rank_list
                assert relevant_document_id in pegasus_rank_list
            except:
                continue

            baseline_rank = baseline_rank_list.index(relevant_document_id)
            ours_rank = pegasus_rank_list.index(relevant_document_id)
            if ours_rank > 100:
                continue
            if baseline_rank - ours_rank < 50:
                continue

            document = document_data[relevant_document_id]["text"]

            generated = mc4_result_data[
                document_data[relevant_document_id]["doc_idx"]
            ][0]

            assert isinstance(generated, str) and isinstance(document, str)

            """
            Novel words finding
            """

            query_wo_stop = list(
                set(
                    [
                        lemmatizer.lemmatize(el).lower()
                        for el in word_tokenize(query)
                        if el not in stopwords
                    ]
                )
            )

            document_wo_stop = list(
                set(
                    [
                        lemmatizer.lemmatize(el).lower()
                        for el in word_tokenize(document)
                        if el not in stopwords
                    ]
                )
            )
            generated_wo_stop = list(
                set(
                    [
                        lemmatizer.lemmatize(el).lower()
                        for el in word_tokenize(generated)
                        if el not in stopwords
                    ]
                )
            )
            novel_gen_query_list = []
            novel_list = []
            for word in generated_wo_stop:
                if word not in document_wo_stop:
                    novel_list.append(word)
                    if word in query_wo_stop:
                        novel_gen_query_list.append(word)

            novel_list = list(set(novel_list))
            if len(novel_list) < 5:
                continue
            novel_list = " ".join(novel_list)

            novel_gen_query_list = list(set(novel_gen_query_list))
            if len(novel_gen_query_list) < 1:
                continue
            novel_gen_query_list = " ".join(novel_gen_query_list)

            final_result.append(
                {
                    "query": query,
                    "document": document,
                    "generated": generated,
                    "baseline_rank": baseline_rank,
                    "pegasus_rank": ours_rank,
                    "novel": novel_list,
                    "novel_q_g_only": novel_gen_query_list,
                    "qid": query_id,
                    "docid": relevant_document_id,
                }
            )
    wb = Workbook()
    ws = wb.active

    for idx, el in enumerate(final_result):
        ws["A" + str(13 * idx + 1)] = "Query"
        ws["A" + str(13 * idx + 1)].font = Font(bold=True)
        ws["A" + str(13 * idx + 2)] = "Doc"
        ws["A" + str(13 * idx + 2)].font = Font(bold=True)
        ws["A" + str(13 * idx + 3)] = "Gen"
        ws["A" + str(13 * idx + 3)].font = Font(bold=True)
        ws["A" + str(13 * idx + 4)] = "Novel"
        ws["A" + str(13 * idx + 4)].font = Font(bold=True)
        ws["A" + str(13 * idx + 5)] = "Contribute"
        ws["A" + str(13 * idx + 5)].font = Font(bold=True)
        ws["A" + str(13 * idx + 6)] = "Qid"
        ws["A" + str(13 * idx + 6)].font = Font(bold=True)
        ws["A" + str(13 * idx + 7)] = "Did"
        ws["A" + str(13 * idx + 7)].font = Font(bold=True)
        ws["A" + str(13 * idx + 8)] = "baseline_rank"
        ws["A" + str(13 * idx + 8)].font = Font(bold=True)
        ws["A" + str(13 * idx + 9)] = "pegasus_rank"
        ws["A" + str(13 * idx + 9)].font = Font(bold=True)
        ws["A" + str(13 * idx + 10)] = "IDX"
        ws["A" + str(13 * idx + 10)].font = Font(bold=True)
        ws["B" + str(13 * idx + 1)] = el["query"]
        ws["B" + str(13 * idx + 2)] = el["document"]
        ws["B" + str(13 * idx + 3)] = el["generated"]
        ws["B" + str(13 * idx + 4)] = el["novel"]
        ws["B" + str(13 * idx + 5)] = el["novel_q_g_only"]
        ws["B" + str(13 * idx + 6)] = el["qid"]
        ws["B" + str(13 * idx + 7)] = el["docid"]
        ws["B" + str(13 * idx + 8)] = el["baseline_rank"]
        ws["B" + str(13 * idx + 9)] = el["pegasus_rank"]
        ws["B" + str(13 * idx + 10)] = idx

    wb.save("effective_sample.xlsx")


def main_0318_legacy():
    baseline_rank, pegasus_rank = get_rank_result()
    print(baseline_rank[list(baseline_rank.keys())[0]])
    score_data = get_score_data()
    query_data = {el[0]: el[1] for el in read_file(query_fname)}
    document_data = {
        el[0]: {"doc_idx": idx, "text": el[1]}
        for idx, el in enumerate(read_file(document_fname))
    }
    qrels_data = read_qrels_file(qrels_fname)
    relevant_docs = get_relevant_doc_list(qrels_data)
    mc4_result_data = read_file(mc4_result_fname)

    final_result = []
    for (query_id, original_score, expanded_score) in score_data:
        if expanded_score - original_score <= 0.4:
            continue
        query = query_data[query_id]
        baseline_rank_list, pegasus_rank_list = (
            baseline_rank[query_id],
            pegasus_rank[query_id],
        )

        relevant_document_list = relevant_docs[query_id]

        for idx, el in enumerate(baseline_rank_list):
            if el in relevant_document_list:
                baseline_first_answer = el
                baseline_top1_idx = idx + 1
                break
        for idx, el in enumerate(pegasus_rank_list):
            if el in relevant_document_list:
                pegasus_first_answer = el
                pegasus_top1_idx = idx + 1
                break
        if baseline_first_answer != pegasus_first_answer:
            continue
        selected_document = document_data[baseline_first_answer]
        generated_text = mc4_result_data[selected_document["doc_idx"]]
        assert len(generated_text) == 1
        generated_text = generated_text[0]
        document = selected_document["text"]

        generated_wo_stop = [
            el.lower()
            for el in word_tokenize(generated_text)
            if el not in stopwords
        ]

        novel_word = []
        for word in generated_wo_stop:
            if word not in document.lower() and word in query.lower():
                novel_word.append(word)
        novel_word = list(set(novel_word))
        novel_word = " ".join(novel_word)
        print(novel_word)
        final_result.append(
            {
                "query": query,
                "document": document,
                "generated": generated_text,
                "baseline_rank": baseline_top1_idx,
                "pegasus_rank": pegasus_top1_idx,
                "baseline_mrr": original_score,
                "pegasus_mrr": expanded_score,
                "novel": novel_word,
                "qid": query_id,
                "docid": baseline_first_answer,
            }
        )
    wb = Workbook()
    ws = wb.active

    for idx, el in enumerate(final_result):
        ws["A" + str(13 * idx + 1)] = "Query"
        ws["A" + str(13 * idx + 1)].font = Font(bold=True)
        ws["A" + str(13 * idx + 2)] = "Doc"
        ws["A" + str(13 * idx + 2)].font = Font(bold=True)
        ws["A" + str(13 * idx + 3)] = "Gen"
        ws["A" + str(13 * idx + 3)].font = Font(bold=True)
        ws["A" + str(13 * idx + 4)] = "Novel"
        ws["A" + str(13 * idx + 4)].font = Font(bold=True)
        ws["A" + str(13 * idx + 5)] = "Qid"
        ws["A" + str(13 * idx + 5)].font = Font(bold=True)
        ws["A" + str(13 * idx + 6)] = "Did"
        ws["A" + str(13 * idx + 6)].font = Font(bold=True)
        ws["A" + str(13 * idx + 7)] = "baseline_rank"
        ws["A" + str(13 * idx + 7)].font = Font(bold=True)
        ws["A" + str(13 * idx + 8)] = "pegasus_rank"
        ws["A" + str(13 * idx + 8)].font = Font(bold=True)
        ws["A" + str(13 * idx + 9)] = "baseline_mrr"
        ws["A" + str(13 * idx + 9)].font = Font(bold=True)
        ws["A" + str(13 * idx + 10)] = "pegasus_mrr"
        ws["A" + str(13 * idx + 10)].font = Font(bold=True)
        ws["A" + str(13 * idx + 11)] = "IDX"
        ws["A" + str(13 * idx + 11)].font = Font(bold=True)
        ws["B" + str(13 * idx + 1)] = el["query"]
        ws["B" + str(13 * idx + 2)] = el["document"]
        ws["B" + str(13 * idx + 3)] = el["generated"]
        ws["B" + str(13 * idx + 4)] = el["novel"]
        ws["B" + str(13 * idx + 5)] = el["qid"]
        ws["B" + str(13 * idx + 6)] = el["docid"]
        ws["B" + str(13 * idx + 7)] = el["baseline_rank"]
        ws["B" + str(13 * idx + 8)] = el["pegasus_rank"]
        ws["B" + str(13 * idx + 9)] = el["baseline_mrr"]
        ws["B" + str(13 * idx + 10)] = el["pegasus_mrr"]
        ws["B" + str(13 * idx + 11)] = idx

    wb.save("effective_sample.xlsx")


def main_legacy():
    dataset = get_matched_dataset()
    counter = [0, 0]
    sample_candidate = []

    for item in dataset:
        query, document, generated = (
            item["query"],
            item["document"],
            item["generated"],
        )
        query_id, doc_id = item["query_id"], item["doc_id"]
        query_wo_stop = [el for el in query if el not in stopwords]
        document_wo_stop = [el for el in document if el not in stopwords]
        generated_wo_stop = [el for el in generated if el not in stopwords]
        lemma_document_wo_stop = [lemmatizer.lemmatize(el) for el in document]
        novel_words = []

        """
        for word in query_wo_stop:
            lemma_word = lemmatizer.lemmatize(word)
            if word not in document_wo_stop and word in generated_wo_stop:
                if word not in " ".join(document):
                    if lemma_word not in " ".join(document):
                        if lemma_word not in ["get", "rid"]:
                            novel_words.append(word)
        """
        for word in generated_wo_stop:
            word = lemmatizer.lemmatize(word)
            if word not in lemma_document_wo_stop:
                novel_words.append(word)
        novel_words = list(set(novel_words))

        if len(novel_words) > 10 and len(document) > 30:
            sample_candidate.append(
                {
                    "query": " ".join(query),
                    "doc": " ".join(document),
                    "generated": " ".join(generated),
                    "novel": novel_words,
                    "qid": query_id,
                    "did": doc_id,
                }
            )
            counter[0] += 1

        else:
            counter[1] += 1
    print(counter)

    write_excel(sample_candidate)


if __name__ == "__main__":
    main()
